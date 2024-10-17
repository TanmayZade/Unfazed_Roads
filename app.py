from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_cors import CORS  # Import CORS
import os
from openpyxl import Workbook, load_workbook
import pandas as pd
import logging

app = Flask(__name__, template_folder='.')
CORS(app)  # Enable CORS for all routes
app.secret_key = 'your_secret_key'  # Necessary for flashing messages

# Create folder to store uploaded images
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Check if Excel files exist, if not, create them
REGISTRATION_FILE = 'registrations_data.xlsx'
REPORT_FILE = 'report_potholes.xlsx'

# Create registrations_data.xlsx if it doesn't exist
if not os.path.exists(REGISTRATION_FILE):
    workbook = Workbook()
    sheet = workbook.active
    # Create header row
    sheet.append(["First Name", "Last Name", "Email", "Phone", "Username", "Password"])
    workbook.save(REGISTRATION_FILE)

# Create report_potholes.xlsx if it doesn't exist
if not os.path.exists(REPORT_FILE):
    workbook = Workbook()
    sheet = workbook.active
    # Create header row
    sheet.append(["Image Path", "Pin Code", "Latitude", "Longitude"])
    workbook.save(REPORT_FILE)

@app.route('/')
def home():
    return render_template('report.html')

@app.route('/register', methods=['POST'])
def register():
    # Get form data
    first_name = request.form['firstName']
    last_name = request.form['lastName']
    email = request.form['email']
    phone = request.form['phone']
    username = request.form['username']
    password = request.form['password']

    # Write the registration data to registrations_data.xlsx
    workbook = load_workbook(REGISTRATION_FILE)
    sheet = workbook.active
    sheet.append([first_name, last_name, email, phone, username, password])
    workbook.save(REGISTRATION_FILE)

        # Check if the username already exists
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row (skipping header)
        if row[4] == username:  # Assuming username is at index 4
            return jsonify({'success': False, 'message': "Username already exists."})

    return jsonify({'success': True, 'message': "Registration successful!"})

@app.route('/report', methods=['POST'])
def report():
    # Handle file upload for pothole reporting
    pincode = request.form['Pin_Code']
    latitude = request.form['latitude']
    longitude = request.form['longitude']

    # Handle file upload
    photo = request.files['photo']
    if photo and photo.filename.endswith('.jpg'):
        image_path = os.path.join(app.config['UPLOAD_FOLDER'], photo.filename)
        photo.save(image_path)
    else:
        return jsonify({'success': False, 'message': "Please upload a JPG file."}), 400

    # Write the pothole report data to report_potholes.xlsx
    workbook = load_workbook(REPORT_FILE)
    sheet = workbook.active
    sheet.append([image_path, pincode, latitude, longitude])
    workbook.save(REPORT_FILE)

    return jsonify({'success': True, 'message': "Pothole report submitted successfully!"})

# Route to get coordinates from the report_potholes.xlsx file
@app.route('/get-coordinates', methods=['GET'])
def get_coordinates():
    if not os.path.exists(REPORT_FILE):
        logging.error("Excel file not found.")
        return jsonify({"error": "Excel file not found"}), 404
    
    # Load the Excel file
    try:
        df = pd.read_excel(REPORT_FILE)
        logging.debug(f"Excel file loaded successfully: {REPORT_FILE}")
    except Exception as e:
        logging.error(f"Error loading Excel file: {str(e)}")
        return jsonify({"error": str(e)}), 500
    
    # Check if required columns exist
    required_columns = ['Latitude', 'Longitude']
    if not all(col in df.columns for col in required_columns):
        missing_cols = [col for col in required_columns if col not in df.columns]
        logging.error(f"Missing columns in Excel file: {missing_cols}")
        return jsonify({"error": f"Excel file must contain columns: {required_columns}"}), 400
    
    # Extract coordinates into a list of dictionaries
    coordinates = df[required_columns].dropna()
    
    # Log the raw coordinates before validation
    logging.debug(f"Raw coordinates extracted: {coordinates.to_dict(orient='records')}")

    # Optionally: Validate the latitude and longitude values
    coordinates = [
        coord for coord in coordinates.to_dict(orient='records') 
        if isinstance(coord['Latitude'], (int, float)) and isinstance(coord['Longitude'], (int, float))
    ]
    
    # Log the validated coordinates
    logging.debug(f"Validated coordinates: {coordinates}")
    
    return jsonify(coordinates)


@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']

    # Load the Excel file
    workbook = load_workbook(REGISTRATION_FILE)
    sheet = workbook.active

    # Check for matching username and password
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] == username and row[5] == password:  # Assuming username is at index 4 and password at index 5
            return jsonify({'success': True, 'message': "Login successful!"})

    return jsonify({'success': False, 'message': "Invalid username or password."})


if __name__ == '__main__':
    app.run(debug=True)
